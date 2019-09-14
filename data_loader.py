import ccxt
from openpyxl import Workbook
import time
from datetime import datetime
import pandas as pd
import numpy as np

# 기존 데이터에 추가로 아래로 이어써지게끔 수정함


# milisecond
msec = 1000
minute = 60 * msec
hold = 30

binance = ccxt.binance({
    'rateLimit': 1200, # set rate limit
    'enableRateLimit': True,
    # 'verbose': True,
})

title_name = ["Date","Open","High","Low","Close","Volume"] # 각 열의 타이틀 설정 ( 인덱스 )

def run():
    # set start date
    start_time = '2019-02-18 00:30:00' # 시작시간만 정해두면 알아서 그 코인 시작 날짜부터 시작하도록 코드 수정해둠
    from_datetime = start_time
    # parsing binance time
    from_timestamp = binance.parse8601(from_datetime)

    workbook = Workbook()
    sheet = workbook.active
    coin_name = 'BCH'
    time_size = '15m'
    file_name_btc = 'data/' + coin_name + '_' + time_size + '.xlsx'
    Original = pd.read_excel(file_name_btc) # 기존 데이터가 없으면 * 표시된 부분은 없애고 돌려야함 *
    Original = np.array(Original) # *
    for i in range(6):
        sheet.cell(row=1, column=i+1).value = title_name[i]
    i = 2

    for org in Original: # *
        for j in range(6): # *
            sheet.cell(row=i,column=j+1).value = org[j] # *
        i = i + 1 # *
    i = i - 1 # *

    sheet.title = time_size

    now = binance.milliseconds()
    k = 1
    ohlcvs = binance.fetch_ohlcv(coin_name + '/USDT', time_size, from_timestamp)  # ('symbol', 'duration', 'start time')
    if(ohlcvs[0][0] > from_timestamp):
        from_timestamp = ohlcvs[0][0]
    while from_timestamp < now:
        try:
            #binance
            ohlcvs = binance.fetch_ohlcv(coin_name + '/USDT', time_size, from_timestamp)  # ('symbol', 'duration', 'start time')
            first = ohlcvs[0][0]
            last = ohlcvs[-1][0]
            date = [datetime.fromtimestamp(d[0]/1000) for d in ohlcvs]
            print('First candle epoch', first, binance.iso8601(first))
            #test_time = from_timestamp - date[0]
            from_timestamp += (len(ohlcvs) * minute * 15) # minute 뒤에 숫자는 몇분 단위로 끊어줄 지 정하는것 하루면 24*60 해야할듯 현재는 15분을 의미함

            opening = [o[1] for o in ohlcvs] # open price
            highest = [h[2] for h in ohlcvs] # highest price
            lowest = [l[3] for l in ohlcvs] # lowest price
            closing = [c[4] for c in ohlcvs] # closing price
            volume = [v[5] for v in ohlcvs] # volume

            row_index = len(ohlcvs)
            print(row_index)

            a = 1

            while 1:
                #sheet.cell(row=k+a, column=1).value = date[a-1]
                #sheet.cell(row=k+a, column=2).value = opening[a-1]
                #sheet.cell(row=k+a, column=3).value = highest[a-1]
                #sheet.cell(row=k+a, column=4).value = lowest[a-1]
                #sheet.cell(row=k+a, column=5).value = closing[a-1]
                #sheet.cell(row=k+a, column=6).value = volume[a-1]

                sheet.cell(row=i + a, column=1).value = date[a - 1] #이미 데이터가 있을때 뒤에다 바로 추가하길 원하면 이거로 바꿔야함
                sheet.cell(row=i + a, column=2).value = opening[a - 1]
                sheet.cell(row=i + a, column=3).value = highest[a - 1]
                sheet.cell(row=i + a, column=4).value = lowest[a - 1]
                sheet.cell(row=i + a, column=5).value = closing[a - 1]
                sheet.cell(row=i + a, column=6).value = volume[a - 1]
                row_index -= 1
                a += 1

                if (row_index <= 0):
                    break

            k += len(closing)
            i += len(closing) # 처음 데이터 받을땐 이 부분은 의미 없음

            workbook.save(filename=file_name_btc)
            time.sleep(0.01)

        except (ccxt.ExchangeError, ccxt.AuthenticationError, ccxt.ExchangeNotAvailable, ccxt.RequestTimeout) as error:

            print('Got an error', type(error).__name__, error.args, ', retrying in', hold, 'seconds...')
            time.sleep(hold)

if __name__ == "__main__":
    run()