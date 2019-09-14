import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from keras.models import Sequential
from keras.layers import LSTM, Dropout, Dense
from time import sleep
from openpyxl import Workbook
import time
from datetime import datetime

coin_name = 'data/ETH_15m'  # 코인 데이터가 있는 디렉토리와 이름
minute = 60
data_time = minute * 15  # 곱하는 숫자는 사용하는 데이터에 분단위에 맞게 하면됨 1시간이면 60 하루면 60 * 24


def normalize_windows(data):  # 정규화 함수
    normalized_data = []
    for window in data:
        _data = [((p / window[0]) - 1) for p in window]
        normalized_data.append(_data)
    return np.array(normalized_data)


def denormalize_windows(org_data, data):  # 역정규화 함수 , 일단 이건 전부 다 입력시에만 가능하도록 근데 이거 아마 안될듯
    denormalized_data = []
    i = 0
    for window in data:
        _data = ((window + 1) * org_data[i][0])
        denormalized_data.append(_data)
        i = i + 1
    return np.array(denormalized_data)


def denormalize_value(org_data, data, location):  # 역정규화 함수인데 전부다 하는것이 아니고 예측된 값 하나만 사용
    normal = org_data[len(org_data) - 1][location]
    return (data + 1) * normal


def normalize_values(data):  # 배열 하나만 정규화 시킬경우
    semi_data = [((value / data[0]) - 1) for value in data]
    return np.array(semi_data)


def prediction_multi_data(model, duration, x_last, new_result, org_result, momentum):  # 앞으로 나올 값 예측할 때 사용하는 함수
    pred_duration = []
    org_pred_value = []
    for i in range(duration):
        pred_value = model.predict(x_last)
        org_pred = denormalize_value(org_result, pred_value, i)
        new_result = new_result[1:, ]
        new_result = np.concatenate((new_result, org_pred), axis=0)
        pred_duration.append(pred_value)
        x_last = normalize_values(new_result)  # 마지막 값들에 대한 정규화 # 정규화된 값을 테스트에 넣을 값으로 변환
        x_last = x_last[:-1]
        x_last = x_last.reshape(1, x_last.shape[0], x_last.shape[1])  # 결과에 맞는 numpy 배열 값으로 변환
        org_pred_value.append(org_pred)
    return np.array(pred_duration), np.array(org_pred_value)


def build_model(inputs):
    model = Sequential()
    model.add(LSTM(100, return_sequences=True, input_shape=(inputs.shape[1], inputs.shape[2])))
    # 위에 크기만큼 숫자 분리(seq_len 만큼) , 뉴런 수를 512로 설정함 이것은 변하는 값
    model.add(Dropout(0.2))
    model.add(LSTM(100, return_sequences=False))  # 총 lstm 층은 두개로 설정함
    model.add(Dropout(0.2))
    model.add(Dense(inputs.shape[2], activation='linear'))
    # model.load_weights("model/" + coin_name.split('/')[1] + '.h5')
    model.load_weights("model/BTC_15m.h5") # BTC, BCH 용
    #model.load_weights("model/ETH_15m.h5")
    model.compile(loss='mae', optimizer='rmsprop', metrics=['accuracy'])
    model.summary()
    return model


def write_excel(Coin, org_pred, coin_name):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'new'
    title_name = ["Date", "Open", "High", "Low", "Close", "Volume"]  # 각 열의 타이틀 설정 ( 인덱스 )
    for i in range(6):
        sheet.cell(row=1, column=i + 1).value = title_name[i]
    i = 2
    org_date = []
    _time = 0
    for coin in Coin:
        for j in range(6):
            sheet.cell(row=i, column=j + 1).value = coin[j]
        _time = time.mktime(coin[0].timetuple())
        i = i + 1
    _time = _time + data_time
    for pred in org_pred:
        _time = datetime.fromtimestamp(_time)
        org_date.append(_time)
        sheet.cell(row=i, column=1).value = _time
        for k in range(5):
            sheet.cell(row=i, column=k + 2).value = pred[k]  # 이부분은 유동적이여야함 , 트레이닝에서 4개를 출력시켰다면
            # 4이고 아니라면 그거보다 적어야함 pred_value 값이랑 같음 나머지 부분은 0으로 채워줘야함
        #sheet.cell(row=i, column=6).value = 0  # 나중에 volume 값 예측이 되는 코드를 통해서 예측이 된다면
        # 그땐 그거에 대해서 추가하면 될듯
        _time = time.mktime(_time.timetuple())
        _time = _time + data_time
        i = i + 1

    workbook.save('result/' + coin_name.split('/')[1] + '_prediction.xlsx')
    return org_date


def run():
    Coin = pd.read_excel(coin_name + '.xlsx')  # 엑셀 째로 불러올경우
    # Coin = pd.read_csv(coin_name + '.csv')

    _Coin = Coin.drop('Date', 1)
    #_Coin = _Coin.drop('Volume', 1)
    _Coin['Volume'] = _Coin['Volume'] + 0.001 # 가끔 volume값에 0이 있는데 0은 정규화에서 오류나기 때문에 추가됨
    _Coin = np.array(_Coin)

    result = []
    seq_len = 100  # Window 사이즈
    sequence_length = seq_len + 1
    for index in range(len(_Coin) - sequence_length):  # 전체 갯수에 따라서 크기 조정
        result.append(_Coin[index: index + sequence_length])

    result = np.array(result)
    org_result = result

    result = normalize_windows(result)
    x_test = result[:, :-1]
    y_test = result[:, -1]

    model = build_model(x_test)  # load 용으로 한번 사용해보자 ( 이건 weight 값만 로드 하고 다른 변수들을 새롭게 하는 것 )
    pred = model.predict(x_test)  # 예측
    #test_pred = denormalize_windows(org_result,pred)
    #print(test_pred)

    x_last = x_test[len(x_test) - 1:, ]  # 마지막 데이터 받아옴
    new_result = org_result[len(org_result) - 1]  # 새로운 window 로 나뉜 마지막 값들
    momentum = [0, 1, 2, 3, 4]  # 앞으로 넘어올 모멘텀 분석에 대한 값 (매수 매도를 정해주는 확실한 지표) 이것을 토대로 예측된 값에 방향을 조금은 수정할 예정
    pred_future, org_pred = prediction_multi_data(model, 96, x_last,
                                                  new_result, org_result,
                                                  momentum)  # 1주일치 예측 이게 잘 되는건진 모르겠음 첫번째 변수는 기간 일주일이면 7
    pred_future = np.reshape(pred_future, (pred_future.shape[0], pred_future.shape[2]))
    pred = np.concatenate((pred, pred_future), axis=0)  # 예측된 부분까지 추가된 예측값(?) 표현이 이상하긴 하지만

    org_pred = org_pred.reshape(org_pred.shape[0], org_pred.shape[2])  # 예측된 값은 정규화된 값이기 때문에 원래 값으로 바꿔줌
    Coin = np.array(Coin)

    pred_date = write_excel(Coin, org_pred, coin_name)  # 엑셀파일로 저장 ( 새로 예측된 부분들 포함 ) 추가로 날짜도 반환

    fig = plt.figure(facecolor='white', figsize=(20, 6))
    ax = fig.add_subplot(111)
    #ax.plot(y_test[:,3], 'r', label='True')
    #ax = fig.add_subplot(212)
    ax.plot(pred_date, org_pred[:, 0],label='Prediction')  # 예측된 결과만 확인할 경우 원래 데이터를 출력할 필요가 없고
    # 그 부분만 보여주면 됨(여기선 Open값을 사용했지만 다른 값을 사용할 꺼면 배열 위치를 바꿔주면됨 )
    #ax.plot(pred[:,3], label='Prediction')
    ax.legend()
    #plt.xlim(len(pred) - 500, len(pred))
    plt.xticks(ha='center')
    plt.tight_layout()
    plt.savefig('result/' + coin_name.split('/')[1] + '.png')
    plt.show()


if __name__ == "__main__":
    run()
